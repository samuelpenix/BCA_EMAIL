import PySimpleGUI as sg
import datetime
import os
import openpyxl
import smtplib
from email.message import EmailMessage
import imghdr
from docx import Document
import comtypes.client

loading = "Waiting"
BCAAmount = 0
ChurchAmount = 0
working_directory = os.getcwd()
layout = [
    [sg.Text("Select BCA template:"), sg.InputText(key="BCA_template_path"), sg.FileBrowse(initial_folder=working_directory, file_types=[("Word Files", "*.docx")])],
    [sg.Text("Select Church template:"), sg.InputText(key="Church_template_path"), sg.FileBrowse(initial_folder=working_directory, file_types=[("Word Files", "*.docx")])],
    [sg.Text("Select Excel file:"), sg.InputText(key="Excel_path"), sg.FileBrowse(initial_folder=working_directory, file_types=[("CSV Files", "*.xlsx")])],
    [sg.Text("Select Receipt destination folder:"), sg.InputText(key='Receipt_folder'), sg.FolderBrowse(initial_folder=working_directory)],
    [sg.Text("Enter password for REPLACE WITH EMAIL:"), sg.InputText(key='password')],
    [sg.Button("Run"),],
    [sg.Text("Status: " + loading, key='load')],
    [sg.Text("BCA emails sent: " + str(BCAAmount), key='BCA')],
    [sg.Text("Church emails sent: " + str(ChurchAmount), key='Church')]
]

window = sg.Window("Donation Email and Receipt Automator", layout, enable_close_attempted_event=True)

while True:
    event, values = window.read()
    if(event == sg.WINDOW_CLOSE_ATTEMPTED_EVENT) and sg.popup_yes_no("Please make sure the program has finished running before exiting. Are you sure you want to exit?") == 'Yes':
        break
    elif event == "Run":
        wb = openpyxl.load_workbook(values['Excel_path'])
        sh1 = wb['Paypal details']
        row = sh1.max_row
        column = sh1.max_column
        BCA = 0
        Church = 0


        def replace_string(filename):
            doc = Document(filename)
            for p in doc.paragraphs:
                if 'DONOR_NAME' or 'DONATION_AMOUNT' or 'CURRENCY_TYPE' or 'DONATION_DATE' in p.text:
                    inline = p.runs
                    for i in range(len(inline)):
                        if 'DONOR_NAME' in inline[i].text:
                            text = inline[i].text.replace('DONOR_NAME', O1.Name.value.lstrip())
                            inline[i].text = text
                        if 'DONATION_AMOUNT' in inline[i].text:
                            text = inline[i].text.replace('DONATION_AMOUNT', str(O1.Amount.value).lstrip())
                            inline[i].text = text
                        if 'CURRENCY_TYPE' in inline[i].text:
                            text = inline[i].text.replace('CURRENCY_TYPE', O1.Currency.value.lstrip())
                            inline[i].text = text
                        if 'DONATION_DATE' in inline[i].text:
                            text = inline[i].text.replace('DONATION_DATE', str(O1.Date.value.date().strftime('%d/%m/%Y')))
                            inline[i].text = text

            doc.save(values['Receipt_folder'] + "\\" + f"{O1.Number.value}-{O1.Name.value}-donation receipt.docx")
            wdFormatPDF = 17
            in_file = os.path.abspath(values['Receipt_folder'] + "\\" + f"{O1.Number.value}-{O1.Name.value}-donation receipt.docx")
            out_file = os.path.abspath(values['Receipt_folder'] + "\\" + f"{O1.Number.value}-{O1.Name.value}-donation receipt.pdf")

            word = comtypes.client.CreateObject('Word.Application')
            doc = word.Documents.Open(in_file)
            doc.SaveAs(out_file, FileFormat=wdFormatPDF)
            doc.Close()
            word.Quit()
            return 1


        class RowObject:
            Number = 0
            Type = 'NA'
            Date = datetime.datetime.now().date()
            Name = 'F L'
            Org = 'NA'
            Email = 'NAME@DOMAIN.COM'
            Currency = 'AAA'
            Amount = 50
            Receipt = 'NO'
            Sent = 'NO'


        O1 = RowObject()

        for i in range(9, row + 1):
            loading = "Running..."
            window['load'].update("Status: " + loading)
            window.refresh()
            for j in range(1, column + 1):
                O1.Number = sh1.cell(i, 1)
                O1.Type = sh1.cell(i, 2)
                O1.Date = sh1.cell(i, 3)
                O1.Name = sh1.cell(i, 4)
                O1.Org = sh1.cell(i, 5)
                O1.Email = sh1.cell(i, 6)
                O1.Currency = sh1.cell(i, 7)
                O1.Amount = sh1.cell(i, 8)
                O1.Receipt = sh1.cell(i, 9)
                O1.Sent = sh1.cell(i, 10)

            if ((O1 is not None)) & (O1.Type.value == "Paypal Chelm Church" or O1.Type.value == "Paypal Chełm Church") & (str(O1.Receipt.value).upper() == "NO") & (str(O1.Sent.value).upper() == "NO"):
                replace_string(values['Church_template_path'])
                ToSend = EmailMessage()
                ToSend['Subject'] = 'Thank you for your donation - BCA Poland to Ukraine'
                ToSend['From'] = 'REPLACE WITH EMAIL'
                ToSend['To'] = O1.Email.value

                ToSend.set_content("Hello " + O1.Name.value.lstrip() + "," +
                                   "\n\nThank you for your donation!" + "\n\n Please see the attached document for your receipt.")
                ToSend.add_alternative("Hello " + O1.Name.value.lstrip() + "," +
                                       "<br /><br />Thank you for your donation!" + "<br /><br /> Please see the attached document for your receipt." + """<br /><br /><div dir="ltr"><div dir="ltr"><div><b><font face="arial narrow, sans-serif" color="#0b5394">Jonasz Skrzypkowski</font></b></div><div><font face="arial narrow, sans-serif"><a href="mailto:akcjabach@gmail.com" target="_blank">akcjabach@gmail.com</a></font></div><div><font face="arial narrow, sans-serif">+48 786 411 053</font></div><span></span><img src="https://ci6.googleusercontent.com/proxy/hiZX4nqfkdYBiLgNQqQcOU4nJ1bFIRLZ7sbwvj-TTP0TsZ5rSiz85ZOM5HaqvyhrmZCPMqt67V9x-qagBhvWlhVRrMD1j9BHYPsWdOBGYS96mYTXggDgXsrzggjNRzYGWxawsM8hfnewfRdAGDEBnTqKHCJk8h2kktbdvxg0imcJqMVcd-x6BhlnVhRkOZ6eWVNB_5OkLy19VEon9w=s0-d-e1-ft#https://docs.google.com/uc?export=download&amp;id=1a5Njcl4mVRSPiU72yi3UrwX70njLITWL&amp;revid=0B8IpApYHS89qR1VVbm9KbXMzaEl4V1cvWjlsMVNEODh4dFpRPQ" width="200" height="50" class="CToWUd"><br><div><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i>„As you did it to one of the least of these my brothers, you did it to me” - Mat. 25:40</i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i><br></i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i><br></i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i><br></i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><br></p></div></div></div>""",
                                       subtype='html')
                with open(values['Receipt_folder'] + "\\" + f"{O1.Number.value}-{O1.Name.value}-donation receipt.pdf", 'rb') as f:
                    file_data = f.read()
                    file_type = imghdr.what(f.name)
                    file_name = f"{O1.Number.value}-{O1.Name.value}-donation receipt.pdf"
                    ToSend.add_attachment(file_data, maintype='application', subtype='pdf', filename=file_name)
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login('REPLACE WITH EMAIL', values['password'])
                server.send_message(ToSend)
                sh1.cell(i, 9, 'YES')
                sh1.cell(i, 10, 'YES')
                wb.save(values['Excel_path'])
                ChurchAmount += 1
                window['Church'].update("Church emails sent: " + str(ChurchAmount))
            elif ((O1 is not None)) & (O1.Type.value == "Paypal BCA") & (str(O1.Receipt.value).upper() == "NO") & (str(O1.Sent.value).upper() == "NO"):
                replace_string(values['BCA_template_path'])
                ToSend = EmailMessage()
                ToSend['Subject'] = 'Thank you for your donation - BCA Poland to Ukraine'
                ToSend['From'] = 'REPALCE WITH EMAIL'
                ToSend['To'] = O1.Email.value

                ToSend.set_content("Hello " + O1.Name.value.lstrip() + "," +
                                   "\n\nThank you for your donation!" + "\n\n Please see the attached document for your receipt.")
                ToSend.add_alternative("Hello " + O1.Name.value.lstrip() + "," +
                                       "<br /><br />Thank you for your donation!" + "<br /><br /> Please see the attached document for your receipt." + """<br /><br /><div dir="ltr"><div dir="ltr"><div><b><font face="arial narrow, sans-serif" color="#0b5394">Jonasz Skrzypkowski</font></b></div><div><font face="arial narrow, sans-serif"><a href="mailto:akcjabach@gmail.com" target="_blank">akcjabach@gmail.com</a></font></div><div><font face="arial narrow, sans-serif">+48 786 411 053</font></div><span></span><img src="https://ci6.googleusercontent.com/proxy/hiZX4nqfkdYBiLgNQqQcOU4nJ1bFIRLZ7sbwvj-TTP0TsZ5rSiz85ZOM5HaqvyhrmZCPMqt67V9x-qagBhvWlhVRrMD1j9BHYPsWdOBGYS96mYTXggDgXsrzggjNRzYGWxawsM8hfnewfRdAGDEBnTqKHCJk8h2kktbdvxg0imcJqMVcd-x6BhlnVhRkOZ6eWVNB_5OkLy19VEon9w=s0-d-e1-ft#https://docs.google.com/uc?export=download&amp;id=1a5Njcl4mVRSPiU72yi3UrwX70njLITWL&amp;revid=0B8IpApYHS89qR1VVbm9KbXMzaEl4V1cvWjlsMVNEODh4dFpRPQ" width="200" height="50" class="CToWUd"><br><div><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i>„As you did it to one of the least of these my brothers, you did it to me” - Mat. 25:40</i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i><br></i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i><br></i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><font face="arial narrow, sans-serif"><i><br></i></font></p><p style="margin:0px;font-stretch:normal;font-size:13px;line-height:normal;color:rgb(0,0,0)"><br></p></div></div></div>""",
                                       subtype='html')
                with open(values['Receipt_folder'] + "\\" + f"{O1.Number.value}-{O1.Name.value}-donation receipt.pdf", 'rb') as f:
                    file_data = f.read()
                    file_type = imghdr.what(f.name)
                    file_name = f"{O1.Number.value}-{O1.Name.value}-donation receipt.pdf"
                    ToSend.add_attachment(file_data, maintype='application', subtype='pdf', filename=file_name)
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login('REPLACE WITH EMAIL', values['password'])
                server.send_message(ToSend)
                sh1.cell(i, 9, 'YES')
                sh1.cell(i, 10, 'YES')
                wb.save(values['Excel_path'])
                BCAAmount += 1
                window['BCA'].update("BCA emails sent: " + str(BCAAmount))
        loading = "Finished"
        window['load'].update("Status: " + loading)
window.close()